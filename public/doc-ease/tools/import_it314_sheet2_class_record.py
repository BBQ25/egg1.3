#!/usr/bin/env python
"""
Import helper: inject docs/2nd Sem 2025 - 2026.xlsx [Sheet2] into E-Record DB.

Target:
- Subject: IT 314 (Functional Programming) lecture
- Sections: BSIT - 3A, BSIT - 3B
- Term: midterm (default)

What gets created/updated:
- class_records.year_level: sets to "3rd Year" if blank for the target class records
- section_grading_configs (per section + term) if missing
- grading_categories (per subject) if missing
- grading_components (per section_config) if missing (with auto weights)
- grading_assessments (per component) if missing
- class_enrollments (roster) for matched students if missing
- grading_assessment_scores for matched students

Run:
  python tools/import_it314_sheet2_class_record.py --dry-run
  python tools/import_it314_sheet2_class_record.py

Notes:
- Students are matched by "Surname, FirstName" string against students.Surname + students.FirstName.
- Non-numeric markers are mapped only for binary (max<=1) assessments:
  - E / excused / with excuse letter => 1
  - A / absent => 0
  Otherwise they are treated as NULL.
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import MySQLdb
import openpyxl


XLSX_PATH = r"docs\2nd Sem 2025 - 2026.xlsx"
SHEET_NAME = "Sheet2"

TARGET_SUBJECT_CODE = "IT 314"
TARGET_SECTIONS = ["BSIT - 3A", "BSIT - 3B"]
TARGET_COURSE = "BSInfoTech"
TARGET_YEAR_LEVEL = "3rd Year"


def norm_name(s: str) -> str:
    s = (s or "").strip().lower()
    s = s.replace("\u00a0", " ")
    # Keep letters/digits/spaces only.
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _strip_suffix_tokens(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"\b(jr|jr\.|sr|sr\.|ii|iii|iv)\b", "", s, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip(" ,")
    return s


def sheet_name_keys(raw: str) -> List[str]:
    raw = (raw or "").strip()
    if not raw or "," not in raw:
        return []
    parts = [p.strip() for p in raw.split(",")]
    if len(parts) < 2:
        return []
    surname = parts[0]
    given = " ".join(parts[1:]).strip()
    if not surname or not given:
        return []

    keys = []
    keys.append(norm_name(surname) + "," + norm_name(given))
    given2 = _strip_suffix_tokens(given)
    if given2 and given2 != given:
        keys.append(norm_name(surname) + "," + norm_name(given2))
    # Dedup while preserving order.
    out = []
    seen = set()
    for k in keys:
        if k and k not in seen:
            seen.add(k)
            out.append(k)
    return out


def parse_header_max(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, dt.datetime):
        # Excel "5" formatted as date can become 1900-01-05.
        if v.year == 1900 and v.month == 1:
            return float(v.day)
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        s = s.replace(",", ".")
        if re.fullmatch(r"\d+(?:\.\d+)?", s):
            return float(s)
    return None


def has_numeric_like_values(ws, start_row: int, end_row: int, col: int) -> bool:
    for r in range(start_row, end_row + 1):
        v = ws.cell(r, col).value
        if v is None or (isinstance(v, str) and v.strip() == ""):
            continue
        if isinstance(v, (int, float)):
            return True
        if isinstance(v, str):
            s = v.strip().replace(",", ".")
            if re.fullmatch(r"\d+(?:\.\d+)?", s):
                return True
    return False


def parse_cell_score(v: Any, max_score: float) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, dt.datetime):
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        up = s.upper()
        # Handle common markers.
        if max_score <= 1.0000001:
            if up in ("E", "EXCUSED", "WITH EXCUSE LETTER", "WITH EXCUSE", "W/ EXCUSE", "WITH EXCUSELETTER"):
                return 1.0
            if up in ("A", "ABSENT", "AB", "a".upper()):
                return 0.0
        if up in ("ABSENT", "A"):
            return 0.0
        # Numeric string.
        s2 = s.replace(",", ".")
        if re.fullmatch(r"\d+(?:\.\d+)?", s2):
            return float(s2)
    return None


def component_type_for(name: str, category: str) -> str:
    s = (name + " " + category).lower()
    if "quiz" in s:
        return "quiz"
    if "exam" in s:
        return "exam"
    if "project" in s:
        return "project"
    if "assignment" in s or "activity" in s:
        return "assignment"
    if "laboratory" in s or "lab" in s:
        return "assignment"
    if "attendance" in s or "recitation" in s or "meet" in s:
        return "participation"
    return "other"


@dataclass
class AssCol:
    col: int
    category: str
    component: str
    assessment: str
    assessment_date: Optional[dt.date]
    max_score: float


def load_sheet_columns(xlsx_path: str) -> Tuple[List[AssCol], int, int, int, int]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {SHEET_NAME}")
    ws = wb[SHEET_NAME]

    # Find used range.
    used_rows, used_cols = set(), set()
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell.value
            if v is None or (isinstance(v, str) and v.strip() == ""):
                continue
            used_rows.add(cell.row)
            used_cols.add(cell.column)
    if not used_rows:
        raise SystemExit("Sheet is empty.")

    min_r, max_r = min(used_rows), max(used_rows)
    min_c, max_c = min(used_cols), max(used_cols)

    # Data start row: first row with a comma-name in first ~10 columns.
    start_row = None
    for r in range(min_r, max_r + 1):
        for c in range(min_c, min(max_c, min_c + 9) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "," in v:
                start_row = r
                break
        if start_row:
            break
    if not start_row:
        raise SystemExit("Could not detect student rows (no comma-names found).")

    header_rows = list(range(min_r, start_row))
    # Determine name column by comma frequency.
    comma_counts: Dict[int, int] = {}
    for c in range(min_c, max_c + 1):
        cnt = 0
        for r in range(start_row, max_r + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "," in v:
                cnt += 1
        if cnt:
            comma_counts[c] = cnt
    if not comma_counts:
        raise SystemExit("Could not detect name column.")
    name_col = max(comma_counts.items(), key=lambda x: x[1])[0]

    def sstr(v: Any) -> str:
        if v is None:
            return ""
        if isinstance(v, dt.datetime):
            return v.date().isoformat()
        return str(v).strip()

    ass_cols: List[AssCol] = []
    for c in range(name_col + 1, max_c + 1):
        # Require at least some data or some header label.
        data_nonempty = False
        for r in range(start_row, max_r + 1):
            v = ws.cell(r, c).value
            if v is None or (isinstance(v, str) and v.strip() == ""):
                continue
            data_nonempty = True
            break

        h1 = sstr(ws.cell(header_rows[0], c).value) if len(header_rows) >= 1 else ""
        h2 = sstr(ws.cell(header_rows[1], c).value) if len(header_rows) >= 2 else ""
        h3v = ws.cell(header_rows[2], c).value if len(header_rows) >= 3 else None
        h3 = sstr(h3v)
        h4v = ws.cell(header_rows[3], c).value if len(header_rows) >= 4 else None

        if not data_nonempty and not (h1 or h2 or h3):
            continue

        # Category: prefer row2 label, else row1 label.
        category = h2 or h1 or "Uncategorized"

        # Normalize patterns like "Attendance (Online Meet)" -> category "Attendance", component "Online Meet".
        m = re.fullmatch(r"\s*attendance\s*\((.+)\)\s*", category, flags=re.I)
        category_inner = m.group(1).strip() if m else None

        # Component naming:
        # If row2 says "Laboratory" but row1 has a specific activity title, prefer row1 as component.
        # Otherwise prefer row2 (e.g., Attendance, Recitation).
        if category_inner:
            category = "Attendance"
            component = category_inner
        elif h2 and norm_name(h2) in {"laboratory", "lab"} and h1:
            component = h1
        else:
            component = h2 or h1 or f"Column {c}"

        # Date (row3) if present.
        assessment_date = None
        if isinstance(h3v, dt.datetime):
            assessment_date = h3v.date()
        elif isinstance(h3v, dt.date):
            assessment_date = h3v

        # Max score: row4, else infer from column values.
        max_score = parse_header_max(h4v)
        if max_score is None:
            maxv = None
            for r in range(start_row, max_r + 1):
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)):
                    maxv = v if maxv is None else max(maxv, v)
                elif isinstance(v, str):
                    vv = v.strip().replace(",", ".")
                    if re.fullmatch(r"\d+(?:\.\d+)?", vv):
                        f = float(vv)
                        maxv = f if maxv is None else max(maxv, f)
            max_score = float(maxv) if maxv is not None else 1.0

        if max_score <= 0:
            max_score = 1.0

        if assessment_date:
            assessment = f"{component} ({assessment_date.isoformat()})"
        else:
            assessment = component

        # Ignore columns that are likely just title headers without real score entries.
        if not has_numeric_like_values(ws, start_row, max_r, c):
            continue

        ass_cols.append(
            AssCol(
                col=c,
                category=category,
                component=component,
                assessment=assessment,
                assessment_date=assessment_date,
                max_score=float(max_score),
            )
        )

    return ass_cols, start_row, name_col, min_r, max_r


def compute_component_weights(ass_cols: List[AssCol]) -> Dict[str, float]:
    # component_name -> weight%
    comp_totals: Dict[str, float] = {}
    for a in ass_cols:
        comp_totals[a.component] = comp_totals.get(a.component, 0.0) + float(a.max_score)
    total = sum(comp_totals.values()) or 1.0

    # Round to 2 decimals, adjust last to sum exactly 100.00.
    items = list(comp_totals.items())
    weights: Dict[str, float] = {}
    running = 0.0
    for i, (comp, pts) in enumerate(items):
        if i == len(items) - 1:
            w = round(100.0 - running, 2)
        else:
            w = round((pts / total) * 100.0, 2)
            running += w
        if w < 0:
            w = 0.0
        weights[comp] = w
    return weights


def main() -> int:
    # Windows console safety: avoid cp1252 crashes for names with non-ASCII characters.
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--term", choices=["midterm", "final"], default="midterm")
    args = ap.parse_args()

    ass_cols, start_row, name_col, min_r, max_r = load_sheet_columns(XLSX_PATH)
    comp_weights = compute_component_weights(ass_cols)

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    ws = wb[SHEET_NAME]

    conn = MySQLdb.connect(host="localhost", user="root", passwd="", db="doc_ease", charset="utf8mb4")
    conn.autocommit(False)
    cur = conn.cursor(MySQLdb.cursors.DictCursor)

    # Resolve subject.
    cur.execute(
        "SELECT id, subject_name, course, academic_year, semester FROM subjects WHERE subject_code=%s AND status='active' LIMIT 1",
        (TARGET_SUBJECT_CODE,),
    )
    subj = cur.fetchone()
    if not subj:
        raise SystemExit(f"Subject not found/active: {TARGET_SUBJECT_CODE}")
    subject_id = int(subj["id"])
    subject_course = (subj.get("course") or "").strip() or TARGET_COURSE

    # Resolve class records per section for this subject (term AY/Sem).
    class_records: Dict[str, Dict[str, Any]] = {}
    for sec in TARGET_SECTIONS:
        cur.execute(
            "SELECT id, teacher_id, academic_year, semester, year_level "
            "FROM class_records "
            "WHERE subject_id=%s AND section=%s AND status='active' "
            "ORDER BY id DESC LIMIT 1",
            (subject_id, sec),
        )
        cr = cur.fetchone()
        if not cr:
            raise SystemExit(f"Missing class_record for {TARGET_SUBJECT_CODE} section {sec}. Assign teacher first.")
        class_records[sec] = cr

    # Normalize year_level on class_records (improves teacher-grading-config compatibility).
    for sec, cr in class_records.items():
        yl = (cr.get("year_level") or "").strip()
        if yl == "":
            cur.execute("UPDATE class_records SET year_level=%s WHERE id=%s", (TARGET_YEAR_LEVEL, int(cr["id"])))

    # Students map (only BSInfoTech 3rd Year).
    cur.execute(
        "SELECT id, StudentNo, Surname, FirstName, Course, Year, Section "
        "FROM students WHERE Course=%s AND Year=%s",
        (TARGET_COURSE, TARGET_YEAR_LEVEL),
    )
    students = cur.fetchall()
    by_key: Dict[str, Dict[str, Any]] = {}
    for st in students:
        surname = st.get("Surname") or ""
        firstname = st.get("FirstName") or ""
        base = norm_name(surname) + "," + norm_name(firstname)
        if base and base not in by_key:
            by_key[base] = st
        v = _strip_suffix_tokens(str(firstname))
        if v:
            k2 = norm_name(surname) + "," + norm_name(v)
            if k2 and k2 not in by_key:
                by_key[k2] = st

    # Ensure grading categories exist (per subject).
    cat_ids: Dict[str, int] = {}
    for a in ass_cols:
        cat = a.category.strip() or "Uncategorized"
        if cat in cat_ids:
            continue
        cur.execute("SELECT id FROM grading_categories WHERE subject_id=%s AND category_name=%s LIMIT 1", (subject_id, cat))
        row = cur.fetchone()
        if row:
            cat_ids[cat] = int(row["id"])
        else:
            cur.execute(
                "INSERT INTO grading_categories (category_name, description, subject_id, default_weight, is_active, created_by) "
                "VALUES (%s, '', %s, 0.00, 1, %s)",
                (cat, subject_id, "system-import"),
            )
            cat_ids[cat] = int(cur.lastrowid)

    # For each section: ensure config + components + assessments.
    cfg_ids: Dict[str, int] = {}  # section -> config_id
    comp_ids: Dict[Tuple[str, str], int] = {}  # (section, component_name) -> id
    assess_ids: Dict[Tuple[str, str], int] = {}  # (section, assessment_name) -> id

    for sec, cr in class_records.items():
        academic_year = (cr.get("academic_year") or "").strip()
        semester = (cr.get("semester") or "").strip()
        year_level = (cr.get("year_level") or "").strip() or TARGET_YEAR_LEVEL

        cur.execute(
            "SELECT id FROM section_grading_configs "
            "WHERE subject_id=%s AND course=%s AND year=%s AND section=%s AND academic_year=%s AND semester=%s AND term=%s "
            "LIMIT 1",
            (subject_id, subject_course, year_level, sec, academic_year, semester, args.term),
        )
        row = cur.fetchone()
        if row:
            cfg_id = int(row["id"])
        else:
            cur.execute(
                "INSERT INTO section_grading_configs "
                "(subject_id, course, year, section, academic_year, semester, term, total_weight, is_active, created_by) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,100.00,1,%s)",
                (subject_id, subject_course, year_level, sec, academic_year, semester, args.term, "system-import"),
            )
            cfg_id = int(cur.lastrowid)
        cfg_ids[sec] = cfg_id

        # Components
        for comp_name in sorted({a.component for a in ass_cols}):
            cur.execute(
                "SELECT id, weight FROM grading_components WHERE section_config_id=%s AND component_name=%s LIMIT 1",
                (cfg_id, comp_name),
            )
            r = cur.fetchone()
            if r:
                comp_id = int(r["id"])
                # If component exists but has 0 weight, set our computed weight (only if empty).
                if float(r.get("weight") or 0) == 0.0:
                    cur.execute("UPDATE grading_components SET weight=%s WHERE id=%s", (comp_weights.get(comp_name, 0.0), comp_id))
            else:
                # Pick category for this component: first matching ass_col's category.
                cat = next((a.category for a in ass_cols if a.component == comp_name), "Uncategorized")
                cat_id = int(cat_ids.get(cat, 0)) if int(cat_ids.get(cat, 0)) > 0 else 0
                ctype = component_type_for(comp_name, cat)
                cur.execute(
                    "INSERT INTO grading_components "
                    "(subject_id, section_config_id, academic_year, semester, course, year, section, category_id, "
                    " component_name, component_code, component_type, weight, is_active, display_order, created_by) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,'',%s,%s,1,0,%s)",
                    (
                        subject_id,
                        cfg_id,
                        academic_year,
                        semester,
                        subject_course,
                        year_level,
                        sec,
                        cat_id,
                        comp_name,
                        ctype,
                        comp_weights.get(comp_name, 0.0),
                        "system-import",
                    ),
                )
                comp_id = int(cur.lastrowid)
            comp_ids[(sec, comp_name)] = comp_id

        # Assessments per column
        for a in ass_cols:
            comp_id = comp_ids[(sec, a.component)]
            cur.execute("SELECT id, max_score FROM grading_assessments WHERE grading_component_id=%s AND name=%s LIMIT 1", (comp_id, a.assessment))
            r = cur.fetchone()
            if r:
                aid = int(r["id"])
                # Fill max_score if empty (0).
                if float(r.get("max_score") or 0) == 0.0:
                    cur.execute("UPDATE grading_assessments SET max_score=%s WHERE id=%s", (float(a.max_score), aid))
            else:
                cur.execute(
                    "INSERT INTO grading_assessments "
                    "(grading_component_id, name, max_score, assessment_date, is_active, display_order, created_by) "
                    "VALUES (%s,%s,%s,%s,1,0,%s)",
                    (comp_id, a.assessment, float(a.max_score), a.assessment_date.isoformat() if a.assessment_date else None, int(cr.get("teacher_id") or 0) or None),
                )
                aid = int(cur.lastrowid)
            assess_ids[(sec, a.assessment)] = aid

    # Ensure roster + insert scores.
    # MySQLdb doesn't have a prepared statement API; we just execute per row.
    scores_inserted = 0
    students_matched = 0
    students_unmatched = 0
    unmatched_names: List[str] = []
    capped = 0

    # Query student id -> existing enrollment per class to avoid repeated upserts (optional).
    # Keep simple: use ON DUPLICATE KEY per row.

    for r in range(start_row, max_r + 1):
        raw_name = ws.cell(r, name_col).value
        if not isinstance(raw_name, str):
            continue
        ks = sheet_name_keys(raw_name)
        if not ks:
            continue
        st = None
        for k in ks:
            st = by_key.get(k)
            if st:
                break
        if not st:
            students_unmatched += 1
            if len(unmatched_names) < 25:
                unmatched_names.append(str(raw_name))
            continue
        students_matched += 1

        sec_letter = (st.get("Section") or "").strip().upper()
        if sec_letter not in ("A", "B"):
            continue
        sec = f"BSIT - 3{sec_letter}"
        if sec not in class_records:
            continue

        cr = class_records[sec]
        class_record_id = int(cr["id"])
        teacher_id = int(cr.get("teacher_id") or 0) or None
        academic_year = (cr.get("academic_year") or "").strip()
        semester = (cr.get("semester") or "").strip()
        year_level = (cr.get("year_level") or "").strip() or TARGET_YEAR_LEVEL

        student_id = int(st["id"])
        today = dt.date.today().isoformat()

        cur.execute(
            "INSERT INTO class_enrollments (class_record_id, student_id, enrollment_date, status, created_by, class_id) "
            "VALUES (%s,%s,%s,'enrolled',%s,%s) "
            "ON DUPLICATE KEY UPDATE status='enrolled', updated_at=CURRENT_TIMESTAMP",
            (class_record_id, student_id, today, teacher_id or 0, class_record_id),
        )

        cfg_id = cfg_ids[sec]

        # Score inserts
        for a in ass_cols:
            aid = assess_ids[(sec, a.assessment)]
            max_score = float(a.max_score)
            val = ws.cell(r, a.col).value
            score = parse_cell_score(val, max_score)
            if score is None:
                continue
            if score < 0:
                score = 0.0
            if score > max_score + 1e-9:
                score = max_score
                capped += 1

            cur.execute(
                "INSERT INTO grading_assessment_scores (assessment_id, student_id, score, recorded_by) "
                "VALUES (%s,%s,%s,%s) "
                "ON DUPLICATE KEY UPDATE score=VALUES(score), recorded_by=VALUES(recorded_by), updated_at=CURRENT_TIMESTAMP",
                (aid, student_id, round(float(score), 2), teacher_id),
            )
            scores_inserted += 1

    if args.dry_run:
        conn.rollback()
    else:
        conn.commit()

    print("Import summary")
    print(f"  subject: {TARGET_SUBJECT_CODE} (id={subject_id})")
    print(f"  term: {args.term}")
    print(f"  assessment_cols: {len(ass_cols)}")
    print("  components:")
    for comp, w in comp_weights.items():
        print(f"    - {comp}: {w:.2f}%")
    print(f"  students matched: {students_matched}")
    print(f"  students unmatched (by name): {students_unmatched}")
    print(f"  scores upserted: {scores_inserted}")
    print(f"  capped_over_max: {capped}")
    if unmatched_names:
        print("  unmatched_examples:")
        for n in unmatched_names:
            print(f"    - {n}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
