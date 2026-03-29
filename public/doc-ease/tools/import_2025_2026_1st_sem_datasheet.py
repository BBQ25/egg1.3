#!/usr/bin/env python
"""
Import helper: 2025-2026 Datasheet (1st Sem) -> Doc-Ease database.

Source workbook format (single sheet, repeated blocks):
- Course row:   [Course Code, Subject Code, Description, Units, Schedule]
- Header row:   [#, StudentNo, Name, ..., MT, FT, AVG, INC]
- Student rows: [n, student_no, "Surname, FirstName Middle", ..., mt, ft, avg, inc]

Targets:
- subjects
- class_records
- schedule_slots
- students
- class_enrollments
- enrollments

Run:
  python tools/import_2025_2026_1st_sem_datasheet.py --dry-run
  python tools/import_2025_2026_1st_sem_datasheet.py
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
from dataclasses import dataclass, field
from typing import Dict, Iterable, List, Optional, Tuple

import MySQLdb
import openpyxl


DEFAULT_XLSX = r"2025-2026 Datasheet.xlsx"
DEFAULT_HOST = "localhost"
DEFAULT_USER = "root"
DEFAULT_PASS = ""
DEFAULT_DB = "doc_ease"
DEFAULT_AY = "2025 - 2026"
DEFAULT_SEM = "1st Semester"
DEFAULT_COURSE = "BSInfoTech"
DEFAULT_CREATED_BY_LABEL = "system-import-1stsem-datasheet"


@dataclass
class StudentGradeRow:
    student_no: str
    full_name: str
    mt_raw: str
    ft_raw: str
    avg_raw: str
    inc_raw: str


@dataclass
class ClassBlock:
    course_code: str
    subject_code: str
    description: str
    units: float
    schedule_raw: str
    students: List[StudentGradeRow] = field(default_factory=list)


def out(msg: str) -> None:
    sys.stdout.write(msg + "\n")


def warn(msg: str) -> None:
    sys.stderr.write("[warn] " + msg + "\n")


def norm_text(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def parse_float(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = norm_text(v).replace(",", ".")
    if not s:
        return None
    if re.fullmatch(r"[+-]?\d+(?:\.\d+)?", s):
        try:
            return float(s)
        except ValueError:
            return None
    return None


def parse_student_name(name_text: str) -> Tuple[str, str, str]:
    """
    Accepts: "Surname, FirstName Middle"
    Returns: (surname, firstname, middlename)
    """
    raw = norm_text(name_text)
    if raw == "":
        return "", "", ""

    if "," in raw:
        left, right = raw.split(",", 1)
        surname = norm_text(left)
        right = norm_text(right)
        parts = right.split()
        if not parts:
            return surname, "", ""
        firstname = parts[0]
        middlename = " ".join(parts[1:]).strip()
        return surname, firstname, middlename

    parts = raw.split()
    if len(parts) == 1:
        return parts[0], "", ""
    if len(parts) == 2:
        return parts[1], parts[0], ""
    return parts[-1], parts[0], " ".join(parts[1:-1]).strip()


def parse_if_code_meta(section_code: str) -> Tuple[str, str]:
    """
    IF-4-A-6 -> ("4th Year", "A")
    """
    section_code = norm_text(section_code).upper()
    m = re.match(r"^IF-(\d+)-([A-Z])-\d+$", section_code)
    if not m:
        return "", ""
    year_num = int(m.group(1))
    sec = m.group(2)

    if year_num == 1:
        year_text = "1st Year"
    elif year_num == 2:
        year_text = "2nd Year"
    elif year_num == 3:
        year_text = "3rd Year"
    else:
        year_text = f"{year_num}th Year"
    return year_text, sec


def parse_days_token(token: str) -> List[int]:
    token = norm_text(token)
    if not token:
        return []
    mapping = {
        "Sun": 0,
        "M": 1,
        "T": 2,
        "W": 3,
        "Th": 4,
        "F": 5,
        "Sat": 6,
    }
    pieces = re.findall(r"Sun|Sat|Th|M|T|W|F", token)
    out_days = []
    for p in pieces:
        if p in mapping:
            out_days.append(mapping[p])
    out_days = sorted(set(out_days))
    return out_days


def to_mysql_time(time12: str) -> Optional[str]:
    t = norm_text(time12).upper().replace(".", "")
    if not t:
        return None
    t = re.sub(r"\s+", " ", t)
    try:
        d = dt.datetime.strptime(t, "%I:%M %p")
        return d.strftime("%H:%M:%S")
    except ValueError:
        return None


def parse_schedule(raw: str) -> Optional[Dict[str, object]]:
    """
    Examples:
      "TBA TF 02:30 PM 03:30 PM"
      "CLab-2 MTh 04:00 PM 05:30 PM"
      "CLab-1 TF 10:00 am 11:00 am"
    """
    s = re.sub(r"\s+", " ", norm_text(raw))
    if s == "":
        return None

    m = re.match(
        r"^(.+?)\s+([A-Za-z]+)\s+(\d{1,2}:\d{2}\s*[APap][Mm])\s+(\d{1,2}:\d{2}\s*[APap][Mm])$",
        s,
    )
    if not m:
        return None

    room = norm_text(m.group(1))
    days_token = norm_text(m.group(2))
    start_12 = norm_text(m.group(3))
    end_12 = norm_text(m.group(4))
    days = parse_days_token(days_token)
    start = to_mysql_time(start_12)
    end = to_mysql_time(end_12)
    if not days or not start or not end:
        return None
    if start >= end:
        return None
    return {
        "room": room,
        "days": days,
        "start_time": start,
        "end_time": end,
    }


def day_short(dow: int) -> str:
    return ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"][dow]


def parse_workbook_blocks(path: str) -> List[ClassBlock]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    blocks: List[ClassBlock] = []
    current: Optional[ClassBlock] = None
    seen_student_header = False

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=8, values_only=True):
        c1 = norm_text(row[0] if len(row) > 0 else "")
        c2 = norm_text(row[1] if len(row) > 1 else "")
        c3 = norm_text(row[2] if len(row) > 2 else "")
        c4 = norm_text(row[3] if len(row) > 3 else "")
        c5 = norm_text(row[4] if len(row) > 4 else "")
        c6 = row[5] if len(row) > 5 else None
        c7 = row[6] if len(row) > 6 else None
        c8 = row[7] if len(row) > 7 else None

        is_block_header = bool(
            re.match(r"^IF-\d+-[A-Z]-\d+$", c1.upper())
            and re.match(r"^[A-Za-z]{2}\s*\d+[A-Za-z]?$", c2)
            and c4 != ""
            and c5 != ""
        )

        if is_block_header:
            if current is not None:
                blocks.append(current)
            units = parse_float(c4)
            current = ClassBlock(
                course_code=c1.upper(),
                subject_code=c2.upper(),
                description=c3,
                units=units if units is not None else 0.0,
                schedule_raw=c5,
                students=[],
            )
            seen_student_header = False
            continue

        if current is None:
            continue

        if c1 == "#" and c2.lower() == "studentno":
            seen_student_header = True
            continue

        if not seen_student_header:
            continue

        # Student rows begin with a numeric index + student no + name.
        if re.fullmatch(r"\d+", c1) and c2 != "" and c3 != "":
            current.students.append(
                StudentGradeRow(
                    student_no=c2,
                    full_name=c3,
                    mt_raw=norm_text(row[4] if len(row) > 4 else ""),
                    ft_raw=norm_text(c6),
                    avg_raw=norm_text(c7),
                    inc_raw=norm_text(c8),
                )
            )

    if current is not None:
        blocks.append(current)

    return blocks


def ensure_reference_values(cur, academic_year: str, semester: str, dry_run: bool) -> None:
    cur.execute(
        "SELECT id FROM academic_years WHERE name = %s LIMIT 1",
        (academic_year,),
    )
    if not cur.fetchone():
        if dry_run:
            out(f"Would insert academic_years: {academic_year}")
        else:
            cur.execute(
                "INSERT INTO academic_years (name, status, sort_order) VALUES (%s, 'active', 0)",
                (academic_year,),
            )

    cur.execute(
        "SELECT id FROM semesters WHERE name = %s LIMIT 1",
        (semester,),
    )
    if not cur.fetchone():
        if dry_run:
            out(f"Would insert semesters: {semester}")
        else:
            cur.execute(
                "INSERT INTO semesters (name, status, sort_order) VALUES (%s, 'active', 0)",
                (semester,),
            )


def build_remarks(mt_raw: str, ft_raw: str, avg_raw: str, inc_raw: str) -> Optional[str]:
    parts: List[str] = []
    mt_num = parse_float(mt_raw)
    ft_num = parse_float(ft_raw)
    avg_num = parse_float(avg_raw)

    if mt_raw and mt_num is None:
        parts.append(f"MT:{mt_raw}")
    if ft_raw and ft_num is None:
        parts.append(f"FT:{ft_raw}")
    if avg_raw and avg_num is None:
        parts.append(f"AVG:{avg_raw}")
    if inc_raw:
        parts.append(f"INC:{inc_raw}")

    if not parts:
        return None
    return "; ".join(parts)


def main() -> int:
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--academic-year", default=DEFAULT_AY)
    ap.add_argument("--semester", default=DEFAULT_SEM)
    ap.add_argument("--course", default=DEFAULT_COURSE)
    ap.add_argument("--created-by-label", default=DEFAULT_CREATED_BY_LABEL)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--db-host", default=DEFAULT_HOST)
    ap.add_argument("--db-user", default=DEFAULT_USER)
    ap.add_argument("--db-pass", default=DEFAULT_PASS)
    ap.add_argument("--db-name", default=DEFAULT_DB)
    args = ap.parse_args()

    dry_run = bool(args.dry_run)
    xlsx_path = args.xlsx
    academic_year = args.academic_year.strip()
    semester = args.semester.strip()
    default_course = args.course.strip()
    created_by_label = args.created_by_label.strip() or DEFAULT_CREATED_BY_LABEL

    out("Import 2025-2026 Datasheet (1st Sem)")
    out(f"Workbook: {xlsx_path}")
    out(f"Target term: {academic_year} | {semester}")
    out("Mode: " + ("DRY RUN" if dry_run else "APPLY"))

    blocks = parse_workbook_blocks(xlsx_path)
    if not blocks:
        out("No class blocks detected. Nothing to import.")
        return 1

    out(f"Detected class blocks: {len(blocks)}")
    out(f"Detected student-grade rows: {sum(len(b.students) for b in blocks)}")

    conn = MySQLdb.connect(
        host=args.db_host,
        user=args.db_user,
        passwd=args.db_pass,
        db=args.db_name,
        charset="utf8mb4",
    )
    conn.autocommit(False)
    cur = conn.cursor()

    counts = {
        "subjects_created": 0,
        "subjects_updated": 0,
        "class_records_created": 0,
        "class_records_reused": 0,
        "slots_created": 0,
        "slots_reactivated": 0,
        "slots_unchanged": 0,
        "students_created": 0,
        "students_updated": 0,
        "class_enroll_inserted": 0,
        "class_enroll_updated": 0,
        "enroll_upserted": 0,
    }
    warnings: List[str] = []

    try:
        # Resolve admin actor.
        admin_id = 1
        cur.execute(
            "SELECT id FROM users WHERE role = 'admin' ORDER BY id ASC LIMIT 1"
        )
        row = cur.fetchone()
        if row and row[0]:
            admin_id = int(row[0])
        out(f"Admin user id: {admin_id}")

        ensure_reference_values(cur, academic_year, semester, dry_run)

        # Cache by subject_code.
        subject_id_by_code: Dict[str, int] = {}

        for b in blocks:
            subject_code = b.subject_code
            is_lab = subject_code.endswith("L")
            subject_type = "Laboratory" if is_lab else "Lecture"
            subject_name = b.description + (" [Laboratory]" if is_lab else "")
            units = float(b.units or (1.0 if is_lab else 2.0))

            if dry_run and subject_code in subject_id_by_code:
                # Simulate previously "created" subjects in dry-run mode.
                continue

            cur.execute(
                "SELECT id FROM subjects WHERE subject_code = %s LIMIT 1",
                (subject_code,),
            )
            subject_row = cur.fetchone()
            if subject_row:
                sid = int(subject_row[0])
                subject_id_by_code[subject_code] = sid
                if dry_run:
                    counts["subjects_updated"] += 1
                else:
                    cur.execute(
                        "UPDATE subjects "
                        "SET subject_name = %s, description = %s, course = %s, "
                        "    academic_year = %s, semester = %s, units = %s, type = %s, status = 'active' "
                        "WHERE id = %s",
                        (
                            subject_name,
                            b.description,
                            default_course,
                            academic_year,
                            semester,
                            units,
                            subject_type,
                            sid,
                        ),
                    )
                    counts["subjects_updated"] += 1
            else:
                if dry_run:
                    sid = -1
                    subject_id_by_code[subject_code] = sid
                    counts["subjects_created"] += 1
                else:
                    cur.execute(
                        "INSERT INTO subjects "
                        "(subject_code, subject_name, description, course, major, academic_year, semester, units, type, status, created_by) "
                        "VALUES (%s, %s, %s, %s, '', %s, %s, %s, %s, 'active', %s)",
                        (
                            subject_code,
                            subject_name,
                            b.description,
                            default_course,
                            academic_year,
                            semester,
                            units,
                            subject_type,
                            admin_id,
                        ),
                    )
                    sid = int(cur.lastrowid)
                    subject_id_by_code[subject_code] = sid
                    counts["subjects_created"] += 1

        # Process each class block (record + roster + grades).
        for b in blocks:
            subject_id = int(subject_id_by_code.get(b.subject_code, 0))
            if subject_id <= 0 and not dry_run:
                warnings.append(
                    f"Skip block {b.course_code}/{b.subject_code}: subject id missing."
                )
                continue

            year_level, section_letter = parse_if_code_meta(b.course_code)

            # class_record lookup
            cur.execute(
                "SELECT id, COALESCE(year_level,''), COALESCE(room_number,'') "
                "FROM class_records "
                "WHERE subject_id = %s AND section = %s AND academic_year = %s AND semester = %s AND status = 'active' "
                "ORDER BY id ASC LIMIT 1",
                (subject_id, b.course_code, academic_year, semester),
            )
            cr = cur.fetchone()
            class_record_id: int
            existing_room_number = ""
            if cr:
                class_record_id = int(cr[0])
                existing_year_level = (cr[1] or "").strip()
                existing_room_number = (cr[2] or "").strip()
                counts["class_records_reused"] += 1
                if not dry_run:
                    # Keep teacher assignment intact; update summary fields only.
                    cur.execute(
                        "UPDATE class_records "
                        "SET year_level = %s, record_type = 'assigned', status = 'active' "
                        "WHERE id = %s",
                        (
                            year_level if year_level else existing_year_level,
                            class_record_id,
                        ),
                    )
            else:
                if dry_run:
                    class_record_id = -1
                    counts["class_records_created"] += 1
                else:
                    try:
                        cur.execute(
                            "INSERT INTO class_records "
                            "(subject_id, teacher_id, record_type, section, academic_year, semester, year_level, created_by, status) "
                            "VALUES (%s, NULL, 'assigned', %s, %s, %s, %s, %s, 'active')",
                            (
                                subject_id,
                                b.course_code,
                                academic_year,
                                semester,
                                year_level,
                                admin_id,
                            ),
                        )
                    except Exception:
                        # Fallback for schemas requiring teacher_id.
                        cur.execute(
                            "INSERT INTO class_records "
                            "(subject_id, teacher_id, record_type, section, academic_year, semester, year_level, created_by, status) "
                            "VALUES (%s, %s, 'assigned', %s, %s, %s, %s, %s, 'active')",
                            (
                                subject_id,
                                admin_id,
                                b.course_code,
                                academic_year,
                                semester,
                                year_level,
                                admin_id,
                            ),
                        )
                    class_record_id = int(cur.lastrowid)
                    counts["class_records_created"] += 1
                    existing_room_number = ""

            # schedule slots
            parsed_sched = parse_schedule(b.schedule_raw)
            if not parsed_sched:
                warnings.append(
                    f"Schedule parse failed for {b.course_code}/{b.subject_code}: {b.schedule_raw}"
                )
            else:
                room = norm_text(parsed_sched["room"])
                for dow in parsed_sched["days"]:
                    st = parsed_sched["start_time"]
                    et = parsed_sched["end_time"]
                    if dry_run:
                        counts["slots_created"] += 1
                        continue

                    cur.execute(
                        "SELECT id, status "
                        "FROM schedule_slots "
                        "WHERE class_record_id = %s AND day_of_week = %s AND start_time = %s AND end_time = %s AND COALESCE(room,'') = %s "
                        "LIMIT 1",
                        (class_record_id, int(dow), st, et, room),
                    )
                    slot = cur.fetchone()
                    if slot:
                        slot_id = int(slot[0])
                        slot_status = norm_text(slot[1]).lower()
                        if slot_status != "active":
                            cur.execute(
                                "UPDATE schedule_slots SET status = 'active', modality = 'face_to_face' WHERE id = %s",
                                (slot_id,),
                            )
                            counts["slots_reactivated"] += 1
                        else:
                            counts["slots_unchanged"] += 1
                    else:
                        cur.execute(
                            "INSERT INTO schedule_slots "
                            "(class_record_id, day_of_week, start_time, end_time, room, modality, status, created_by) "
                            "VALUES (%s, %s, %s, %s, %s, 'face_to_face', 'active', %s)",
                            (class_record_id, int(dow), st, et, room, admin_id),
                        )
                        counts["slots_created"] += 1

                if not dry_run:
                    # Refresh class_records.schedule summary from active slots.
                    cur.execute(
                        "SELECT day_of_week, start_time, end_time, COALESCE(room,'') "
                        "FROM schedule_slots "
                        "WHERE class_record_id = %s AND status = 'active' "
                        "ORDER BY day_of_week ASC, start_time ASC, id ASC",
                        (class_record_id,),
                    )
                    slot_rows = cur.fetchall()
                    parts = []
                    room_set = set()
                    for dow, st, et, room_val in slot_rows:
                        st_txt = str(st)[:5]
                        et_txt = str(et)[:5]
                        room_txt = norm_text(room_val)
                        p = f"{day_short(int(dow))} {st_txt}-{et_txt}"
                        if room_txt:
                            p += f" ({room_txt})"
                            room_set.add(room_txt)
                        parts.append(p)
                    summary = "; ".join(parts)
                    cur.execute(
                        "UPDATE class_records SET schedule = %s WHERE id = %s",
                        (summary, class_record_id),
                    )

                    if len(room_set) == 1 and existing_room_number == "":
                        only_room = list(room_set)[0]
                        cur.execute(
                            "UPDATE class_records SET room_number = %s WHERE id = %s AND COALESCE(room_number,'') = ''",
                            (only_room, class_record_id),
                        )

            # Build/ensure section catalog mapping.
            if not dry_run:
                cur.execute(
                    "INSERT INTO class_sections (code, description, status, source) "
                    "VALUES (%s, '', 'active', 'class_records') "
                    "ON DUPLICATE KEY UPDATE status = 'active', updated_at = CURRENT_TIMESTAMP",
                    (b.course_code,),
                )
                cur.execute(
                    "SELECT id FROM class_sections WHERE code = %s LIMIT 1",
                    (b.course_code,),
                )
                class_section_id = int(cur.fetchone()[0])
                cur.execute(
                    "INSERT IGNORE INTO class_section_subjects (class_section_id, subject_id) VALUES (%s, %s)",
                    (class_section_id, subject_id),
                )

            # Student + roster + queue enrollment.
            for sr in b.students:
                surname, firstname, middlename = parse_student_name(sr.full_name)
                if not sr.student_no or not surname or not firstname:
                    warnings.append(
                        f"Skip malformed student row in {b.course_code}/{b.subject_code}: "
                        f"{sr.student_no} | {sr.full_name}"
                    )
                    continue

                # student upsert
                cur.execute(
                    "SELECT id FROM students WHERE StudentNo = %s LIMIT 1",
                    (sr.student_no,),
                )
                st = cur.fetchone()
                if st:
                    student_id = int(st[0])
                    if dry_run:
                        counts["students_updated"] += 1
                    else:
                        cur.execute(
                            "UPDATE students "
                            "SET Surname = %s, FirstName = %s, MiddleName = %s, "
                            "    Course = %s, Year = %s, Section = %s "
                            "WHERE id = %s",
                            (
                                surname,
                                firstname,
                                middlename,
                                default_course,
                                year_level if year_level else "1st Year",
                                section_letter if section_letter else "",
                                student_id,
                            ),
                        )
                        counts["students_updated"] += 1
                else:
                    if dry_run:
                        student_id = -1
                        counts["students_created"] += 1
                    else:
                        cur.execute(
                            "INSERT INTO students "
                            "(user_id, StudentNo, Surname, FirstName, MiddleName, Course, Major, Status, Year, Section, created_by) "
                            "VALUES (NULL, %s, %s, %s, %s, %s, '', 'Continuing', %s, %s, %s)",
                            (
                                sr.student_no,
                                surname,
                                firstname,
                                middlename,
                                default_course,
                                year_level if year_level else "1st Year",
                                section_letter if section_letter else "",
                                admin_id,
                            ),
                        )
                        student_id = int(cur.lastrowid)
                        counts["students_created"] += 1

                # profile section mapping
                if not dry_run:
                    profile_year = year_level if year_level else "1st Year"
                    profile_section = section_letter if section_letter else ""
                    if profile_section:
                        label = f"{default_course} - {profile_year} - {profile_section}"
                        cur.execute(
                            "INSERT INTO profile_sections (course, year_level, section_code, label, status, source) "
                            "VALUES (%s, %s, %s, %s, 'active', 'students') "
                            "ON DUPLICATE KEY UPDATE label = VALUES(label), status = 'active', updated_at = CURRENT_TIMESTAMP",
                            (default_course, profile_year, profile_section, label),
                        )
                        cur.execute(
                            "SELECT id FROM profile_sections WHERE course = %s AND year_level = %s AND section_code = %s LIMIT 1",
                            (default_course, profile_year, profile_section),
                        )
                        ps_row = cur.fetchone()
                        if ps_row:
                            profile_section_id = int(ps_row[0])
                            cur.execute(
                                "INSERT IGNORE INTO profile_section_subjects (profile_section_id, subject_id) VALUES (%s, %s)",
                                (profile_section_id, subject_id),
                            )

                avg_val = parse_float(sr.avg_raw)
                grade_val = None if avg_val is None else round(float(avg_val), 2)
                remarks = build_remarks(sr.mt_raw, sr.ft_raw, sr.avg_raw, sr.inc_raw)
                today = dt.date.today().isoformat()

                # class_enrollments upsert (actual roster + grade)
                if dry_run:
                    # We do not query existence for speed in dry-run; count as upserted.
                    counts["class_enroll_inserted"] += 1
                else:
                    cur.execute(
                        "INSERT INTO class_enrollments "
                        "(class_record_id, student_id, enrollment_date, status, grade, remarks, created_by, class_id) "
                        "VALUES (%s, %s, %s, 'enrolled', %s, %s, %s, %s) "
                        "ON DUPLICATE KEY UPDATE "
                        "  enrollment_date = VALUES(enrollment_date), "
                        "  status = 'enrolled', "
                        "  grade = VALUES(grade), "
                        "  remarks = VALUES(remarks), "
                        "  class_id = VALUES(class_id), "
                        "  updated_at = CURRENT_TIMESTAMP",
                        (
                            class_record_id,
                            student_id,
                            today,
                            grade_val,
                            remarks,
                            admin_id,
                            class_record_id,
                        ),
                    )
                    if cur.rowcount == 1:
                        counts["class_enroll_inserted"] += 1
                    elif cur.rowcount == 2:
                        counts["class_enroll_updated"] += 1
                    else:
                        # MySQL can return 0 when duplicate row has identical values.
                        counts["class_enroll_updated"] += 1

                # enrollments upsert (queue parity)
                if dry_run:
                    counts["enroll_upserted"] += 1
                else:
                    cur.execute(
                        "INSERT INTO enrollments "
                        "(student_no, subject_id, academic_year, semester, section, status, created_by) "
                        "VALUES (%s, %s, %s, %s, %s, 'Claimed', %s) "
                        "ON DUPLICATE KEY UPDATE "
                        "  section = VALUES(section), "
                        "  status = 'Claimed', "
                        "  created_by = VALUES(created_by)",
                        (
                            sr.student_no,
                            subject_id,
                            academic_year,
                            semester,
                            b.course_code,
                            created_by_label,
                        ),
                    )
                    counts["enroll_upserted"] += 1

        if dry_run:
            conn.rollback()
        else:
            conn.commit()
    except Exception as exc:
        conn.rollback()
        sys.stderr.write("FAILED: " + str(exc) + "\n")
        return 1
    finally:
        cur.close()
        conn.close()

    out("")
    out("Summary:")
    for k in [
        "subjects_created",
        "subjects_updated",
        "class_records_created",
        "class_records_reused",
        "slots_created",
        "slots_reactivated",
        "slots_unchanged",
        "students_created",
        "students_updated",
        "class_enroll_inserted",
        "class_enroll_updated",
        "enroll_upserted",
    ]:
        out(f"  {k}: {counts[k]}")

    if warnings:
        out("")
        out(f"Warnings: {len(warnings)}")
        for w in warnings[:30]:
            warn(w)
        if len(warnings) > 30:
            warn(f"... {len(warnings) - 30} more warnings suppressed")
    else:
        out("Warnings: 0")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
